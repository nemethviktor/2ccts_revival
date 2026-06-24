#!/usr/bin/env python3
""" Roster Workbook Gap Analysis Engine
Copyright (C) 2026 V. Nemeth

This module provides native Excel compilation pipelines to track OpenTTD 
vehicle availability lifespans and identify regional model gaps.
"""

import os
from typing import List, Dict, Any, Optional
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.table import Table
import warnings
import re

# Silence openpyxl warnings
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")


def extract_excel_table(sheet: Worksheet) -> pd.DataFrame:
    """ Detects and extracts a structured Excel Table from a worksheet. """
    tables: Dict[str, Table] = sheet.tables
    if not tables:
        data_values: List[List[Any]] = list(sheet.values)
        if not data_values:
            return pd.DataFrame()
        return pd.DataFrame(data=data_values[1:], columns=data_values[0])

    table_obj: Table = next(iter(tables.values()))
    table_range: str = table_obj.ref

    data_rows: List[List[Any]] = []
    for row_cells in sheet[table_range]:
        data_rows.append([cell.value for cell in row_cells])

    df_table: pd.DataFrame = pd.DataFrame(
        data=data_rows[1:], columns=data_rows[0])
    return df_table


def load_workbook_data(workbook_path: str) -> pd.DataFrame:
    """ Loads sheets from the source workbook and joins them via vehicle IDs. """
    wb = load_workbook(filename=workbook_path, data_only=True)

    if "properties" not in wb.sheetnames:
        raise KeyError("Required worksheet 'properties' missing from target workbook.")
    df_prop: pd.DataFrame = extract_excel_table(sheet=wb["properties"])

    if "roster" not in wb.sheetnames:
        raise KeyError("Required worksheet 'roster' missing from target workbook.")
    roster_sheet: Worksheet = wb["roster"]
    roster_values: List[List[Any]] = list(roster_sheet.values)
    df_rost: pd.DataFrame = pd.DataFrame(
        data=roster_values[1:], columns=roster_values[0])

    if "control" not in wb.sheetnames:
        raise KeyError("Required worksheet 'control' missing from target workbook.")
    df_control: pd.DataFrame = extract_excel_table(sheet=wb["control"])

    wb.close()

    df_prop = df_prop.dropna(subset=["VEHIDCODE"])
    df_rost = df_rost.dropna(subset=["VEHIDCODE"])
    df_control = df_control.dropna(subset=["VEHIDCODE"])

    df_prop["VEHIDCODE"] = df_prop["VEHIDCODE"].astype(str).str.strip()
    df_rost["VEHIDCODE"] = df_rost["VEHIDCODE"].astype(str).str.strip()
    df_control["VEHIDCODE"] = df_control["VEHIDCODE"].astype(str).str.strip()

    df_merged: pd.DataFrame = pd.merge(
        left=df_prop, right=df_rost, on="VEHIDCODE", suffixes=("_prop", "_rost")
    )
    df_merged = pd.merge(
        left=df_merged, right=df_control, on="VEHIDCODE", suffixes=("", "_control")
    )

    return df_merged


def calculate_timeline(df_merged: pd.DataFrame) -> pd.DataFrame:
    """ Simulates year-by-year vehicle model availability windows. """
    start_year: int = 1840
    end_year_limit: int = 2050
    years_range: np.ndarray = np.arange(start_year, end_year_limit + 1)

    region_columns: List[str] = [
        "AFRICA", "ASIA", "SOUTHERN_EUROPE", "EASTERN_EUROPE",
        "WESTERN_EUROPE", "NORTHERN_EUROPE", "NORTH_AMERICA",
        "SOUTH_AMERICA", "OCEANIA"
    ]

    timeline_records: List[Dict[str, Any]] = []

    for _, row in df_merged.iterrows():
        if str(object=row.get("EXCLUDE_prop")).strip().lower() == "true":
            continue
        if str(object=row.get("EXCLUDE_rost")).strip().lower() == "true":
            continue

        intro_year: int = int(float(row["INTRODUCTION_YEAR"]))
        model_life_raw: str = str(object=row["MODEL_LIFE"]).strip()

        retire_early = 0 if row.get("IS_WAGON_OR_COACH_control") == True else 20
        if model_life_raw == "VEHICLE_NEVER_EXPIRES" or model_life_raw == "0":
            expiry_year: int = end_year_limit
        else:
            try:
                expiry_year = intro_year + int(float(model_life_raw)) - retire_early
            except ValueError:
                expiry_year = intro_year + 45 - retire_early

        role_category: str = str(object=row["ROLE_rost"])

        for region in region_columns:
            region_flag = row.get(region)
            if region_flag is True or str(object=region_flag).strip().lower() in ["true", "1", "1.0"]:
                # Vectorized generation of records per active region to eliminate the inner year loop
                valid_years = years_range[(years_range >= intro_year) & (years_range <= expiry_year)]
                for current_year in valid_years:
                    timeline_records.append({
                        "Year": current_year,
                        "Region": region,
                        "Category": role_category,
                        "Vehicle": row["VEHIDCODE"]
                    })

    return pd.DataFrame(data=timeline_records)


def generate_visualization_matrix(df_timeline: pd.DataFrame, output_path: str) -> None:
    """Generates independent, region-centric dashboard grid JPG files highly optimized via Vectorization."""
    
    if df_timeline.empty:
        print("No records available to plot.")
        return

    # 1. Compute aggregate dense counts efficiently
    df_counts = df_timeline.groupby(
        by=["Region", "Category", "Year"]
    ).size().reset_index(name="Available_Count")

    target_categories: List[str] = sorted(df_timeline["Category"].unique().tolist())
    
    region_columns: List[str] = [
        "AFRICA", "ASIA", "SOUTHERN_EUROPE", "EASTERN_EUROPE",
        "WESTERN_EUROPE", "NORTHERN_EUROPE", "NORTH_AMERICA",
        "SOUTH_AMERICA", "OCEANIA"
    ]
    simulation_years: np.ndarray = np.arange(1840, 2041)

    # 2. Build a comprehensive MultiIndex representing every possible combination
    full_index = pd.MultiIndex.from_product(
        [region_columns, target_categories, simulation_years],
        names=["Region", "Category", "Year"]
    )
    
    # Reindex fills all missing combinations with 0 instantly (No more nested loops!)
    df_dense = (df_counts.set_index(["Region", "Category", "Year"])
                .reindex(full_index, fill_value=0)
                .reset_index())

    # 3. Iterate through regions to compile plots
    for region in region_columns:
        safe_region_name: str = re.sub(pattern=r"[^a-zA-Z0-9_\-]", repl="_", string=region).lower()
        
        # Isolate this region's data entirely
        df_region = df_dense[df_dense["Region"] == region]

        fig, axes = plt.subplots(nrows=4, ncols=5, figsize=(24, 18), sharex=True, sharey=False)
        flat_axes = axes.flatten()

        for index, category in enumerate(target_categories):
            current_ax = flat_axes[index]

            # Fast linear slice using vectorized matching
            df_series = df_region[df_region["Category"] == category]

            current_ax.plot(
                df_series["Year"],
                df_series["Available_Count"],
                color="#e31a1c",
                linewidth=2.0
            )

            current_ax.set_title(label=category, fontsize=10, fontweight="bold")
            current_ax.grid(visible=True, linestyle=":", alpha=0.6)
            current_ax.tick_params(labelbottom=True)
            current_ax.set_xlabel(xlabel="Year", fontsize=8)
            current_ax.set_ylabel(ylabel="Models", fontsize=8)

            if category == "Wagon":
                ymax = 100
            elif category.startswith("Coach"):
                ymax = 30
            else:
                ymax = 25
            current_ax.set_ylim([0, ymax])

        # Clean up empty subplots
        for empty_index in range(len(target_categories), len(flat_axes)):
            fig.delaxes(ax=flat_axes[empty_index])

        clean_title: str = region.replace("_", " ").title()
        fig.suptitle(
            t=f"Regional Availability Gap Matrix - {clean_title} (All Roles)\n"
              f"Please note that Y-axis maxima aren't identical within a region but they are identical across categories for other regions.", 
            fontsize=20, fontweight="bold", y=0.98
        )

        filename: str = f"gap_analysis_region_{safe_region_name}.jpg"
        plt.savefig(
            fname=os.path.join(output_path, filename),
            format="jpg",
            dpi=72,
            pil_kwargs={"quality": 92, "optimize": True}
        )
        plt.close(fig=fig)
        print(f"---- Exported cross-platform validation asset: {filename}")


def main() -> None:
    """ Execution entry point for the Excel layout compiler. """
    script_dir = os.path.dirname(os.path.abspath(__file__))
    project_root = os.path.dirname(script_dir)
    excel_path = os.path.join(script_dir, 'vehicle_report.xlsx')
    output_path = os.path.normpath(os.path.join(project_root, 'docs', 'gap_analysis'))

    if not os.path.exists(path=excel_path):
        print(f"[Error] Target workbook '{excel_path}' not found.")
        return

    print(f"--- Extracting structured layout data from '{excel_path}'...")
    df_merged_specs: pd.DataFrame = load_workbook_data(workbook_path=excel_path)
    
    print("--- Simulating vehicle lifespans...")
    df_calculated_timeline: pd.DataFrame = calculate_timeline(df_merged=df_merged_specs)
    
    print("--- Generating high-resolution matrix dashboards...")
    generate_visualization_matrix(df_timeline=df_calculated_timeline, output_path=output_path)


if __name__ == "__main__":
    main()