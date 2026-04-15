from typing import Literal

import pandas as pd
import openpyxl
import os
import math
import warnings
from pandas.api.types import is_number
import re

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")


def load_master_data(excel_path):
    print("--- Loading and Merging Excel Sheets with Note Extraction ---")
    # Load raw data for pandas
    sheets = pd.read_excel(excel_path, sheet_name=None)

    # Load workbook for openpyxl (to get notes/comments)
    wb = openpyxl.load_workbook(excel_path)

    # This will store notes as: notes_lookup[VEHIDCODE][COLUMN_NAME] = "Note Text"
    notes_lookup = {}

    def extract_notes(sheet_name, dataframe):
        if sheet_name not in wb.sheetnames:
            return
        ws = wb[sheet_name]

        # Identify where 'VEHIDCODE' is in this specific sheet
        try:
            vehid_col_idx = list(dataframe.columns).index('VEHIDCODE')
        except ValueError:
            return  # Skip sheets without the ID key

        # Iterate Excel rows (skip header)
        for r_idx, row in enumerate(ws.iter_rows(min_row=2), start=0):
            # Get the VEHIDCODE for this row to use as the primary key
            # openpyxl is 1-indexed, so c_idx+1
            veh_id = ws.cell(row=r_idx+2, column=vehid_col_idx+1).value
            if veh_id is None:
                continue

            veh_id = veh_id.lower()

            if veh_id not in notes_lookup:
                notes_lookup[veh_id] = {}

            for c_idx, cell in enumerate(row):
                if cell.comment:
                    col_name = dataframe.columns[c_idx]
                    # Clean the note text (removing Excel's 'Author:' prefix if present)
                    clean_note = cell.comment.text.split(':')[-1].strip()
                    notes_lookup[veh_id][col_name] = f"{cell.value} -- {clean_note}"

    # 1. Start with the 'control' sheet as the base
    df_master = sheets['control']
    extract_notes('control', sheets['control'])

    # 2. List of sheets that provide extra vehicle properties
    data_sheets = [
        'properties', 'flags', 'track_types',
        'graphics_properties'
    ]

    text_columns = ['COUNTRY', 'COUNTRY_CODE', 'ITEM',
                    'NAME', 'VEHIDCODE', 'CARGODEF', 'WEB']

    for sheet_name in data_sheets:
        if sheet_name in sheets:
            current_sheet = sheets[sheet_name]

            # Extract notes from this sheet before merging
            extract_notes(sheet_name, current_sheet)

            # --- Your existing merge logic ---
            cols_to_fix = [
                c for c in text_columns if c in current_sheet.columns]
            for col in cols_to_fix:
                current_sheet[col] = current_sheet[col].astype(
                    str).replace('nan', '')

            overlapping_cols = [c for c in current_sheet.columns
                                if c in df_master.columns and c != 'VEHIDCODE']

            df_master = df_master.drop(columns=overlapping_cols)
            df_master = pd.merge(df_master, current_sheet,
                                 on='VEHIDCODE', how='left')

    # 3. Load the Lookups and Copyright
    df_cost_lookup = sheets['cost_lookup'].set_index('COST_CAT').fillna(0)

    df_copyright = sheets['copyright_text']
    copyright_txt = ""
    if not df_copyright.empty:
        copyright_txt = str(df_copyright.iloc[0, 0])
    elif len(df_copyright.columns) > 0 and "Unnamed" not in str(df_copyright.columns[0]):
        copyright_txt = str(df_copyright.columns[0])

    return df_master, df_cost_lookup, copyright_txt, notes_lookup


def is_true(val) -> bool:
    """ Checks if a value evals to true (ie is a string that says so, or 1, or just True)"""
    return (val == True or str(val).upper() == 'TRUE') or (val == 1)


def get_badges(row: pd.Series) -> str:
    # example: badges: ["type/bus", "power/diesel", "flag/flag_CC", "usage/city"];
    # ok that's from a bus-nml but docu is s.it and can't find a better one.
    vehidcode = row['VEHIDCODE'].lower()
    power = row['ENGINE_CLASS'].lower()
    if (is_true(row['IS_TURBINE'])):
        power = "turbine"
    flag = row['COUNTRY_CODE'].upper()
    badges = []

    if vehidcode.startswith('mtro') or vehidcode.startswith('singlemtro'):
        power = 'metro'

    if power != "":
        if power == 'electric':
            if \
                    is_true(row['TRACK_TYPE_STANDARD_GAUGE_RAILTYPE_25KV']) or \
                    is_true(row['TRACK_TYPE_STANDARD_GAUGE_RAILTYPE_15KV']) or \
                    is_true(row['TRACK_TYPE_NARROW_GAUGE_RAILTYPE_25KV']) or \
                    is_true(row['TRACK_TYPE_NARROW_GAUGE_RAILTYPE_15KV']) or \
                    is_true(row['TRACK_TYPE_BROAD_GAUGE_RAILTYPE_25KV']) or \
                    is_true(row['TRACK_TYPE_BROAD_GAUGE_RAILTYPE_15KV']):
                power = 'electric/ac'
            elif \
                    is_true(row['TRACK_TYPE_STANDARD_GAUGE_RAILTYPE_3KV']) or \
                    is_true(row['TRACK_TYPE_STANDARD_GAUGE_RAILTYPE_1500V']) or \
                    is_true(row['TRACK_TYPE_NARROW_GAUGE_RAILTYPE_3KV']) or \
                    is_true(row['TRACK_TYPE_NARROW_GAUGE_RAILTYPE_1500V']) or \
                    is_true(row['TRACK_TYPE_BROAD_GAUGE_RAILTYPE_3KV']) or \
                    is_true(row['TRACK_TYPE_BROAD_GAUGE_RAILTYPE_1500V']):
                power = 'electric/dc'
            else:
                power = 'electric'

        badges.append(f"power/{power}")
    if flag != "":
        badges.append(f"flag/{flag}")

    badge_string = '", "'.join(badges) if badges else ""

    return f"""\n\tbadges: ["{badge_string}"];\n"""


# --- Newton-Raphson Emulation (Matches NML SQRT) ---
def nml_sqrt(value) -> float:
    """
    Emulates the NML 'SQRTESTIMATE' macro using the Newton-Raphson method.

    This function replicates the specific mathematical approximation used in
    legacy NML property files. It performs 3 iterations to converge on a
    square root value, matching the precision (or lack thereof) of the
    original Newton-Raphson implementation in the game's compilation process.

    Args:
        value (float/int): The radicand to calculate the square root for.

    Returns:
        float: The approximated square root after 3 iterations.
               Returns 0 if the input is less than or equal to 0.
    """
    val = float(value)
    if val <= 0:
        return 0
    # Initial guess is VALUE/40 per your eval string
    guess = val / 40
    for _ in range(3):
        # guess - ((guess^2 - val) / (0.1 + 2*guess))
        guess = guess - ((guess**2) - val) / (0.1 + 2 * guess)
    return guess


def calculate_nml_cost(row: pd.Series, m, is_running_cost=False) -> float:
    """
    Universal cost calculator for Purchase and Running Cost properties.

    This function handles the logic branching between 'Engine/MU' and
    'Coach/Wagon' cost formulas. It pulls multipliers (P1-P7 or R1-R6)
    from the cost_lookup sheet and vehicle data from the properties sheet.

    Logic Branches:
    1. COACH/WAGON: Uses 'PURCHASECOSTNONENGINEVALUE' logic where capacity
       is added linearly rather than using a square root complexity factor.
    2. ENGINES/MU: Uses a complexity-based formula where Power and Capacity
       are processed through the nml_sqrt function.

    Args:
        row (pd.Series): A row from the merged 'master_df' containing vehicle specs.
        m (pd.Series): The row from 'cost_lookup' corresponding to the vehicle category.
        is_running_cost (bool): If True, calculates 'running_cost_factor'.
                                If False, calculates 'cost_factor'.

    Returns:
        float: The final cost factor rounded to 5 decimal places.
    """
    # 1. Setup raw variables and handle empty cells (NaN)
    m = m.fillna(0)
    W = float(row.get('WEIGHT', 0))
    P = float(row.get('POWER', 0))
    S = float(row.get('SPEED', 0))
    TE = float(row.get('TE_COEFFICIENT', 0))
    C = float(row.get('HEAD_CAPACITY', 0))

    # Emulate the 3-iteration Newton SQRT
    sqrt_S = nml_sqrt(S)
    sqrt_P = nml_sqrt(P)

    # 2. Branch Logic based on your Macro Definitions
    if row['COST_CAT'] in ['COACH', 'WAGON']:
        if not is_running_cost:
            # Matches PURCHASECOSTNONENGINEVALUE(SCALAR, WFACTOR, SFACTOR, CAPFACTOR)
            # Logic: SCALAR * (WFACTOR * WEIGHT + SFACTOR * SQRT(SPEED) + CAPFACTOR * CAPACITY)
            # For Coach: 0.5 * (0.1 * 28 + 0.05 * 8.94 + 1 * 60)
            inner_math = (m.P2 * W) + (m.P3 * sqrt_S) + (m.P5 * C)
            return round(m.P1 * inner_math, 5)
        else:
            # Matches RUNNINGCOSTNONENGINEVALUE(SCALAR, SFACTOR, CAPFACTOR)
            # Logic: SCALAR * (SFACTOR * SQRT(SPEED) + CAPACITY)
            inner_math = (m.R2 * sqrt_S) + C
            return round(m.R1 * inner_math, 5)

    else:
        # Standard Engine/MU Logic (Power & Capacity use SQRT)
        sqrt_C = nml_sqrt(C)
        if not is_running_cost:
            base = (m.P2 * W) + (m.P3 * sqrt_S) + \
                (m.P4 * sqrt_P) + (m.P5 * sqrt_C) + (m.P7 * TE)
            if row['COST_CAT'] in ['DMU', 'EMU', 'METRO', 'MAGLEVMU']:
                base += (m.P6 * nml_sqrt(row.get('WAGON_POWER', 0)))
            return round(m.P1 * base, 5)
        else:
            base = (m.R2 * sqrt_S) + (m.R3 * sqrt_P) + \
                (m.R4 * sqrt_C) + (m.R6 * TE)
            if row['COST_CAT'] in ['DMU', 'EMU', 'METRO', 'MAGLEVMU']:
                base += (m.R5 * nml_sqrt(row.get('WAGON_POWER', 0)))
            return round(max(m.R1 * base, 1.0), 5)


def get_climates(row: pd.Series) -> str:
    # 1. Define the Mapping (Excel Text -> NML Variable)
    region_map = {
        "AFRICA": "param_region_africa",
        "NORTH_AMERICA": "param_region_north_america",
        "SOUTH_AMERICA": "param_region_south_america",
        "ASIA": "param_region_asia",
        "NORTHERN_EUROPE": "param_region_northern_europe",
        "EASTERN_EUROPE": "param_region_eastern_europe",
        "WESTERN_EUROPE": "param_region_western_europe",
        "SOUTHERN_EUROPE": "param_region_southern_europe",
        "OCEANIA": "param_region_oceania",
        "NO_REGION": "0",
        "NO_CONCEPT": "1",
        "IS_CONCEPT": "param_concept"
    }

    # 2. Define Complex Macro Mapping
    # These reference the basic variables defined above
    ALL_EUROPE = f"({region_map['NORTHERN_EUROPE']} || {region_map['EASTERN_EUROPE']} || {region_map['SOUTHERN_EUROPE']} || {region_map['WESTERN_EUROPE']})"
    ALL_AMERICA = f"({region_map['NORTH_AMERICA']} || {region_map['SOUTH_AMERICA']})"

    region_map.update({
        "ALL_EUROPE": ALL_EUROPE,
        "ALL_AMERICA": ALL_AMERICA,
        "REGIONTYPE1": f"({ALL_EUROPE} || {region_map['ASIA']} || {region_map['AFRICA']})",
        "REGIONTYPE2": f"({ALL_AMERICA} || {region_map['ASIA']} || {region_map['OCEANIA']} || {region_map['AFRICA']})",
        "REGIONTYPE3": f"({region_map['ASIA']} || {region_map['OCEANIA']})",
        "ALL_REGION":  f"({region_map['AFRICA']} || {ALL_AMERICA} || {region_map['ASIA']} || {ALL_EUROPE} || {region_map['OCEANIA']})"
    })

    # 3. Extract values from Excel Row (stripping whitespace to be safe)
    # Adjust column names if they differ in your master_df
    r1_raw = str(row.get('REGION1', 'NO_REGION')).strip()
    r2_raw = str(row.get('REGION2', 'NO_REGION')).strip()
    # Assuming CONCEPT is in Region3 column
    c_raw = str(row.get('REGION3', 'NO_CONCEPT')).strip()

    # 4. Perform the "Magic" Lookup
    # .get(key, default) ensures that if Excel has a typo, it defaults to NO_REGION/NO_CONCEPT
    region1 = region_map.get(r1_raw, "0")
    region2 = region_map.get(r2_raw, "0")
    concept = region_map.get(c_raw, "1")

    # 5. Build the final string
    return f"climates_available: (({region1} || {region2}) && {concept}) ? ALL_CLIMATES : NO_CLIMATE;"


def get_cargo_definitions(row: pd.Series) -> str:
    """
    Translates the CARGODEF column into full NML cargo property strings.
    This replaces the need for an external cargorefits.pnml file.
    """

    # Define common reusable lists to keep the dictionary clean
    PASS_MAIL_VAL = "CC_PASSENGERS, CC_MAIL, CC_ARMOURED"
    EXPRESS_REF = "CC_PIECE_GOODS, CC_EXPRESS, CC_REFRIGERATED"

    NO_NONREFITTABLE = "non_refittable_cargo_classes: 0;"

    # Standard cargo list used in your DMU example
    STANDARD_ALLOW = "GOOD, PAPR, FOOD, RUBR, FRUT, WATR, TOYS, BATT, SWET, COLA, BUBL, PLST, FZDR, BEER, BDMT, BRCK, CERA, CERE, COPR, DYES, ENSP, FERT, FICR, FISH, FMSP, GLAS, JAVA, MILK, MNSP, PETR, PLAS, RCYC, RFPR, VPTS, WDPR, WOOL, URAN"
    STANDARD_DISALLOW = "PASS, MAIL, TOUR, COAL, OIL_, LVST, GRAI, WOOD, IORE, STEL, VALU, WHEA, GOLD, MAIZ, CORE, DIAM, SUGR, TOFF, CTCD, AORE, CLAY, CMNT, GRVL, LIME, OLSD, POTA, SAND, SCMT, SGBT, SGCN, SULP, VEHI, YETI, YETY"

    cargo_map = {
        "NONE": (
            f"// cargodeftype: NONE;\n{" "*8}"
            f"refittable_cargo_classes: bitmask(NO_CARGO_CLASS);\n{" "*8}"
        ),
        # Pax & Mail
        "PASSENGERS": (
            f"// cargodeftype: PASSENGERS;\n{" "*8}"
            f"refittable_cargo_classes: bitmask({PASS_MAIL_VAL});\n{" "*8}"
            f"{NO_NONREFITTABLE}"
        ),
        # Metro does not allow it
        "PASSENGERS_ONLY": (
            f"// cargodeftype: PASSENGERS_ONLY;\n{" "*8}"
            f"refittable_cargo_classes: bitmask(CC_PASSENGERS);\n{" "*8}"
            f"{NO_NONREFITTABLE}\n{" "*8}"
        ),
        "MAIL_ONLY": (
            f"// cargodeftype: MAIL_ONLY;\n{" "*8}"
            f"refittable_cargo_classes: bitmask(CC_MAIL);\n{" "*8}"
            f"{NO_NONREFITTABLE}\n{" "*8}"
        ),
        # All other
        "BOXCAR": (
            f"// cargodeftype: BOXCAR;\n{" "*8}"
            f"refittable_cargo_classes: bitmask(CC_PIECE_GOODS, CC_EXPRESS, CC_ARMOURED);\n{" "*8}"
            f"{NO_NONREFITTABLE}\n{" "*8}"
            f"cargo_allow_refit: [LVST, GOOD, GRAI, VALU, PAPR, WHEA, FOOD, GOLD, FRUT, MAIZ, DIAM, SUGR, TOYS, BATT, SWET, BUBL, FZDR, BEER, BDMT, BRCK, CERA, CERE, COPR, ENSP, FERT, FISH, FMSP, GLAS, JAVA, MNSP, OLSD, POTA, RCYC, SGBT, SGCN, SULP, VEHI, VPTS, WOOL, URAN];\n{" "*8}"
            f"cargo_disallow_refit: [PASS, MAIL, TOUR, COAL, OIL_, WOOD, IORE, STEL, RUBR, CORE, WATR, TOFF, COLA, CTCD, PLST, AORE, CLAY, CMNT, DYES, FICR, GRVL, LIME, MILK, PETR, PLAS, RFPR, SAND, SCMT, WDPR, YETI, YETY];\n{" "*8}"
        ),
        "CARTRANSPORTER": (
            f"// cargodeftype: CARTRANSPORTER;\n{" "*8}"
            f"refittable_cargo_classes: bitmask();\n{" "*8}"
            f"{NO_NONREFITTABLE}\n{" "*8}"
            f"cargo_allow_refit: [PASS,VEHI,ENSP,FMSP];\n{" "*8}"
            f"cargo_disallow_refit: [];\n{" "*8}"
        ),
        "CENTERBEAM": (
            f"// cargodeftype: CENTERBEAM;\n{" "*8}"
            f"refittable_cargo_classes: bitmask(CC_PIECE_GOODS, CC_EXPRESS);\n{" "*8}"
            f"{NO_NONREFITTABLE}\n{" "*8}"
            f"cargo_allow_refit: [GOOD, STEL, PAPR, TOYS, BATT, SWET, BUBL, FZDR, BDMT, BRCK, CERA, COPR, ENSP, FERT, FMSP, GLAS, JAVA, MNSP, RCYC, VPTS, WDPR, WOOL, URAN];\n{" "*8}"
            f"cargo_disallow_refit: [PASS, MAIL, TOUR, COAL, OIL_, LVST, GRAI, WOOD, IORE, VALU, WHEA, FOOD, GOLD, RUBR, FRUT, MAIZ, CORE, WATR, DIAM, SUGR, TOFF, COLA, CTCD, PLST, AORE, BEER, CERE, CLAY, CMNT, DYES, FICR, FISH, GRVL, LIME, MILK, OLSD, PETR, PLAS, POTA, RFPR, SAND, SCMT, SGBT, SGCN, SULP, VEHI, YETI, YETY];\n{" "*8}"
        ),
        "CONTAINER": (
            f"// cargodeftype: CONTAINER;\n{" "*8}"
            f"refittable_cargo_classes: bitmask({EXPRESS_REF}, CC_POTABLE);\n{" "*8}"
            f"{NO_NONREFITTABLE}\n{" "*8}"
            f"cargo_allow_refit: [{STANDARD_ALLOW}];\n{" "*8}"
            f"cargo_disallow_refit: [{STANDARD_DISALLOW}];\n{" "*8}"
        ),
        "DOUBLECONTAINER": (
            f"// cargodeftype: DOUBLECONTAINER;\n{" "*8}"
            f"refittable_cargo_classes: bitmask(CC_PIECE_GOODS, CC_EXPRESS, CC_REFRIGERATED);\n{" "*8}"
            f"{NO_NONREFITTABLE}\n{" "*8}"
            f"cargo_allow_refit: [GOOD, PAPR, FOOD, FRUT, TOYS, BATT, SWET, BUBL, FZDR, BDMT, BRCK, CERA, CERE, COPR, ENSP, FERT, FICR, FISH, FMSP, GLAS, JAVA, MNSP, RCYC, VPTS, WDPR, WOOL, URAN];\n{" "*8}"
            f"cargo_disallow_refit: [PASS, MAIL, TOUR, COAL, OIL_, LVST, GRAI, WOOD, IORE, STEL, VALU, WHEA, GOLD, RUBR, MAIZ, CORE, WATR, DIAM, SUGR, TOFF, COLA, CTCD, PLST, AORE, BEER, CLAY, CMNT, DYES, GRVL, LIME, MILK, OLSD, PETR, PLAS, POTA, RFPR, SAND, SCMT, SGBT, SGCN, SULP, VEHI, YETI, YETY];\n{" "*8}"
        ),
        "FLAT_WAGON": (
            f"// cargodeftype: FLAT_WAGON;\n{" "*8}"
            f"refittable_cargo_classes: bitmask(CC_PIECE_GOODS);\n{" "*8}"
            f"non_refittable_cargo_classes: bitmask(CC_FLATBED);\n{" "*8}"
            f"cargo_allow_refit: [GOOD, WOOD, STEL, TOYS, BATT, SWET, BUBL, FZDR, BDMT, BRCK, CERA, COPR, ENSP, FICR, FMSP, JAVA, MNSP, VPTS, WDPR, YETI, YETY];\n{" "*8}"
            f"cargo_disallow_refit: [PASS, MAIL, TOUR, COAL, OIL_, LVST, GRAI, IORE, VALU, PAPR, WHEA, FOOD, GOLD, RUBR, FRUT, MAIZ, CORE, WATR, DIAM, SUGR, TOFF, COLA, CTCD, PLST, AORE, BEER, CERE, CLAY, CMNT, DYES, FERT, FISH, GLAS, GRVL, LIME, MILK, OLSD, PETR, PLAS, POTA, RCYC, RFPR, SAND, SCMT, SGBT, SGCN, SULP, VEHI, WOOL, URAN];\n{" "*8}"
        ),
        "GONDOLA": (
            f"// cargodeftype: GONDOLA;\n{" "*8}"
            f"refittable_cargo_classes: bitmask(CC_OPEN_BULK, CC_COVERED_BULK);\n{" "*8}"
            f"{NO_NONREFITTABLE}\n{" "*8}"
            f"cargo_allow_refit: [COAL, GRAI, WOOD, IORE, WHEA, MAIZ, CORE, SUGR, TOFF, CTCD, AORE, CERE, CLAY, CMNT, GRVL, LIME, POTA, SAND, SCMT, WDPR];\n{" "*8}"
            f"cargo_disallow_refit: [PASS, MAIL, TOUR, OIL_, LVST, GOOD, STEL, VALU, PAPR, FOOD, GOLD, RUBR, FRUT, WATR, DIAM, TOYS, BATT, SWET, COLA, BUBL, PLST, FZDR, BEER, BDMT, BRCK, CERA, COPR, DYES, ENSP, FERT, FICR, FISH, FMSP, GLAS, JAVA, MILK, MNSP, OLSD, PETR, PLAS, RCYC, RFPR, SGBT, SGCN, SULP, VEHI, VPTS, WOOL, URAN, YETI, YETY];\n{" "*8}"
        ),
        "GOODS_RAILBUS": (
            f"// cargodeftype: GOODS_RAILBUS;\n{" "*8}"
            f"refittable_cargo_classes: bitmask({EXPRESS_REF});\n{" "*8}"
            f"{NO_NONREFITTABLE}\n{" "*8}"
            f"cargo_allow_refit:[MAIL, GOOD, VALU, GOLD, DIAM];\n{" "*8}"
            f"cargo_disallow_refit: [PASS, TOUR, COAL, OIL_, LVST, GRAI, WOOD, IORE, STEL, PAPR, WHEA, FOOD, RUBR, FRUT, MAIZ, CORE, WATR, SUGR, TOYS, BATT, SWET, TOFF, COLA, CTCD, BUBL, PLST, FZDR, AORE, BEER, BDMT, BRCK, CERA, CERE, CLAY, CMNT, COPR, DYES, ENSP, FERT, FICR, FISH, FMSP, GLAS, GRVL, JAVA, LIME, MILK, MNSP, OLSD, PETR, PLAS, POTA, RCYC, RFPR, SAND, SCMT, SGBT, SGCN, SULP, VEHI, VPTS, WDPR, WOOL, URAN, YETI, YETY];\n{" "*8}"
        ),
        "HEAVYFLAT": (
            f"// cargodeftype: HEAVYFLAT;\n{" "*8}"
            f"refittable_cargo_classes: bitmask(CC_PIECE_GOODS, CC_FLATBED);\n{" "*8}"
            f"{NO_NONREFITTABLE}\n{" "*8}"
            f"cargo_allow_refit: [GOOD, STEL, TOYS, BATT, SWET, BUBL, FZDR, BDMT, BRCK, CERA, COPR, ENSP, FMSP, GLAS, JAVA, MNSP, VEHI, VPTS, YETI, YETY];\n{" "*8}"
            f"cargo_disallow_refit: [PASS, MAIL, TOUR, COAL, OIL_, LVST, GRAI, WOOD, IORE, VALU, PAPR, WHEA, FOOD, GOLD, RUBR, FRUT, MAIZ, CORE, WATR, DIAM, SUGR, TOFF, COLA, CTCD, PLST, AORE, BEER, CERE, CLAY, CMNT, DYES, FERT, FICR, FISH, GRVL, LIME, MILK, OLSD, PETR, PLAS, POTA, RCYC, RFPR, SAND, SCMT, SGBT, SGCN, SULP, WDPR, WOOL, URAN];\n{" "*8}"
        ),
        "HOPPER": (
            f"// cargodeftype: HOPPER;\n{" "*8}"
            f"refittable_cargo_classes: bitmask(CC_OPEN_BULK, CC_COVERED_BULK);\n{" "*8}"
            f"non_refittable_cargo_classes: bitmask(CC_WEIRD);\n{" "*8}"
            f"cargo_allow_refit: [COAL, IORE, CORE, AORE, CLAY, CMNT, GRVL, LIME, SAND, SGBT];\n{" "*8}"
            f"cargo_disallow_refit: [PASS, MAIL, TOUR, OIL_, LVST, GOOD, GRAI, WOOD, STEL, VALU, PAPR, WHEA, FOOD, GOLD, RUBR, FRUT, MAIZ, WATR, DIAM, SUGR, TOYS, BATT, SWET, TOFF, COLA, CTCD, BUBL, PLST, FZDR, BEER, BDMT, BRCK, CERA, CERE, COPR, DYES, ENSP, FERT, FICR, FISH, FMSP, GLAS, JAVA, MILK, MNSP, OLSD, PETR, PLAS, POTA, RCYC, RFPR, SCMT, SGCN, SULP, VEHI, VPTS, WDPR, WOOL, URAN, YETI, YETY];\n{" "*8}"
        ),
        "OPEN_WAGON": (
            f"// cargodeftype: OPEN_WAGON;\n{" "*8}"
            f"refittable_cargo_classes: bitmask(CC_OPEN_BULK);\n{" "*8}"
            f"{NO_NONREFITTABLE}\n{" "*8}"
            f"cargo_allow_refit: [COAL, GRAI, WOOD, IORE, WHEA, FRUT, MAIZ, CORE, SUGR, TOFF, CTCD, BUBL, AORE, CERE, CLAY, CMNT, GRVL, LIME, OLSD, POTA, SAND, SCMT];\n{" "*8}"
            f"cargo_disallow_refit: [PASS, MAIL, TOUR, OIL_, LVST, GOOD, STEL, VALU, PAPR, FOOD, GOLD, RUBR, WATR, DIAM, TOYS, BATT, SWET, COLA, PLST, FZDR, BEER, BDMT, BRCK, CERA, COPR, DYES, ENSP, FERT, FICR, FISH, FMSP, GLAS, JAVA, MILK, MNSP, PETR, PLAS, RCYC, RFPR, SGBT, SGCN, SULP, VEHI, VPTS, WDPR, WOOL, URAN, YETI, YETY];\n{" "*8}"
        ),
        "SUPERHEAVY": (
            f"// cargodeftype: SUPERHEAVY;\n{" "*8}"
            f"refittable_cargo_classes: bitmask(CC_PIECE_GOODS, CC_FLATBED);\n{" "*8}"
            f"{NO_NONREFITTABLE}\n{" "*8}"
            f"cargo_allow_refit: [GOOD, VEHI];\n{" "*8}"
            f"cargo_disallow_refit: [PASS, MAIL, TOUR, COAL, OIL_, LVST, GRAI, WOOD, IORE, STEL, VALU, PAPR, WHEA, FOOD, GOLD, RUBR, FRUT, MAIZ, CORE, WATR, DIAM, SUGR, TOYS, BATT, SWET, TOFF, COLA, CTCD, BUBL, PLST, FZDR, AORE, BEER, BDMT, BRCK, CERA, CERE, CLAY, CMNT, COPR, DYES, ENSP, FERT, FICR, FISH, FMSP, GLAS, GRVL, JAVA, LIME, MILK, MNSP, OLSD, PETR, PLAS, POTA, RCYC, RFPR, SAND, SCMT, SGBT, SGCN, SULP, VPTS, WDPR, WOOL, URAN, YETI, YETY];\n{" "*8}"
        ),
        "SILO": (
            f"// cargodeftype: SILO;\n{" "*8}"
            f"refittable_cargo_classes: bitmask(CC_POWDER_BULK);\n{" "*8}"
            f"{NO_NONREFITTABLE}\n{" "*8}"
            f"cargo_allow_refit: [GRAI, WHEA, MAIZ, SUGR, CERE, OLSD, POTA, SULP, URAN];\n{" "*8}"
            f"cargo_disallow_refit: [PASS, MAIL, TOUR, COAL, OIL_, LVST, GOOD, WOOD, IORE, STEL, VALU, PAPR, FOOD, GOLD, RUBR, FRUT, CORE, WATR, DIAM, TOYS, BATT, SWET, TOFF, COLA, CTCD, BUBL, PLST, FZDR, AORE, BEER, BDMT, BRCK, CERA, CLAY, CMNT, COPR, DYES, ENSP, FERT, FICR, FISH, FMSP, GLAS, GRVL, JAVA, LIME, MILK, MNSP, PETR, PLAS, RCYC, RFPR, SAND, SCMT, SGBT, SGCN, VEHI, VPTS, WDPR, WOOL, YETI, YETY];\n{" "*8}"
        ),
        "TANKER": (
            f"// cargodeftype: TANKER;\n{" "*8}"
            f"refittable_cargo_classes: bitmask(CC_LIQUID_BULK, CC_GAS_BULK);\n{" "*8}"
            f"{NO_NONREFITTABLE}\n{" "*8}"
            f"cargo_allow_refit: [OIL_, GOOD, RUBR, WATR, COLA, PLST, BEER, DYES, MILK, PETR, PLAS, RFPR];\n{" "*8}"
            f"cargo_disallow_refit: [PASS, MAIL, TOUR, COAL, LVST, GRAI, WOOD, IORE, STEL, VALU, PAPR, WHEA, FOOD, GOLD, FRUT, MAIZ, CORE, DIAM, SUGR, TOYS, BATT, SWET, TOFF, CTCD, BUBL, FZDR, AORE, BDMT, BRCK, CERA, CERE, CLAY, CMNT, COPR, ENSP, FERT, FICR, FISH, FMSP, GLAS, GRVL, JAVA, LIME, MNSP, OLSD, POTA, RCYC, SAND, SCMT, SGBT, SGCN, SULP, VEHI, VPTS, WDPR, WOOL, URAN, YETI, YETY];\n{" "*8}"
        ),

    }

    raw_val = str(row.get('CARGODEF', 'NONE')).strip().upper()

    # Return the mapped string, or a fallback if the category is missing
    return cargo_map.get(raw_val, cargo_map["NONE"])


def parse_cargo_definitions(pnml_path):
    """
    Parses cargorefits.pnml and returns a dictionary:
    {'CARGODEF_CONTAINER': 'refittable_cargo_classes: ...', ...}
    """
    cargo_dict = {}
    if not os.path.exists(pnml_path):
        print(f"Warning: {pnml_path} not found. Using raw macro names.")
        return cargo_dict

    # Regex captures the macro name and the entire body until the next #define or EOF
    pattern = re.compile(
        r'#define\s+(CARGODEF_[A-Z0-9_]+)\s+(.*?)(?=\s*#define|$)', re.DOTALL)

    with open(pnml_path, 'r', encoding='utf-8') as f:
        content = f.read()
        # Strip comments to prevent matching commented-out definitions
        content = re.sub(r'//.*', '', content)
        content = re.sub(r'/\*.*?\*/', '', content, flags=re.DOTALL)

        matches = pattern.findall(content)
        for name, body in matches:
            # Clean up whitespace/newlines within the definition body
            cargo_dict[name.strip()] = " ".join(body.split())

    return cargo_dict


def get_expanded_engine_capacity_switch(row: pd.Series) -> str:
    # We use \\ to produce a single literal \ in the output
    # We use {{ }} to produce literal { } in the NML code
    cap = row['HEAD_CAPACITY']
    vehid_lcase = row['VEHIDCODE'].lower()

    return f"""
    switch(FEAT_TRAINS, SELF, switch_{vehid_lcase}_capacity_engine, cargo_classes) {{ \\
        bitmask(CC_MAIL): {cap}/2; \\
        bitmask(CC_ARMOURED): {cap}/4; \\
        {cap}; \\
    }}\n\n"""


def get_expanded_wagon_capacity_switch(row: pd.Series) -> str:
    # We use \\ to produce a single literal \ in the output
    # We use {{ }} to produce literal { } in the NML code
    cap = row['WAGON_CAPACITY']
    vehid_lcase = row['VEHIDCODE'].lower()

    return f"""
    switch(FEAT_TRAINS, SELF, switch_{vehid_lcase}_capacity_wagon, cargo_classes) {{ \\
        bitmask(CC_MAIL): {cap}/2; \\
        bitmask(CC_ARMOURED): {cap}/4; \\
        {cap}; \\
    }}\n\n"""


# --- Main Generation Function ---


def generate_unified_items():
    print("--- Starting Unified Item Generation ---")
    script_dir = os.path.dirname(os.path.abspath(__file__))
    project_root = os.path.dirname(script_dir)
    excel_path = os.path.join(script_dir, 'vehicle_report.xlsx')
    AIR_DRAG_COEFFICIENT = 0
    BITMASK_VEHICLE_INFO = 0
    CARGO_AGE_PERIOD = 185
    POWER_PER_WAGON = 0
    REFIT_COST = 0
    RELIABILITY_DECAY = 20
    RETIRE_EARLY = 20
    SPRITE_ID = "SPRITE_ID_NEW_TRAIN"

    # 1. Load Data
    df_master, df_cost_lookup, copyright_text, notes_lookup = load_master_data(
        excel_path=excel_path)
    for _, row in df_master.iterrows():
        if pd.isna(row['VEHIDCODE']):
            continue
        VEHIDCODE_lcase = row['VEHIDCODE'].lower()
        veh_notes: dict = notes_lookup.get(VEHIDCODE_lcase, {})
        TEMPLATE_ID = row['TEMPLATE_ID']
        TEMPLATE_AMENDMENT_CODE = row['TEMPLATE_AMENDMENT_CODE']

        TEMPLATE_ID_FULL = f"{TEMPLATE_ID}{TEMPLATE_AMENDMENT_CODE}"

        RUNNING_COST_BASE = f"RUNNING_COST_{'ELECTRIC' if row['ENGINE_CLASS'] == 'MAGLEV' else row['ENGINE_CLASS']}"

        # Fetch Multipliers
        category = str(row['COST_CAT']).strip()
        m = df_cost_lookup.loc[category]

        # Calculate Costs for main item
        p_cost = calculate_nml_cost(row, m, is_running_cost=False)
        r_cost = calculate_nml_cost(row, m, is_running_cost=True)

        # 2. Tracks Logic
        tracks = [col.replace('TRACK_TYPE_', '') for col in df_master.columns
                  if col.startswith('TRACK_TYPE_') and (is_true(row[col]))]
        track_logic = f"[{', '.join(tracks)}]"

        # 3. Logic: Misc Flags
        flags = [col.replace('MISC_FLAGS_', '') for col in df_master.columns
                 if col.startswith('MISC_FLAGS_') and (is_true(row[col]))]
        misc_logic = f"bitmask({', '.join(flags)})"
        v1 = f"VISUAL_EFFECT_{row['VISUAL_EFFECT_1']}" if str(
            row['VISUAL_EFFECT_1']) != "0" else "0"

        # Loading Speed
        ls_val = int(row['LOADINGSPEED_VALUE'])
        ls_logic = f"isUltraSpeed ? 255 : (param_loadingspeed == 0) ? {ls_val}/2 : (param_loadingspeed == 2) ? {ls_val}*2 : {ls_val}"

        purchase_cargo_capacity = row['PURCHASE_CARGO_CAPACITY'] if is_number(
            row['PURCHASE_CARGO_CAPACITY']) else None

        # Separate from the further below because this is for powered/unpowered livery overrides only
        graphics_switch_visual_effect_and_powered_position = f"visual_effect_and_powered: switch_{VEHIDCODE_lcase}_visual_effect_and_powered_position;" if TEMPLATE_ID_FULL in [
            'TPL_02D'] else None

        # Ie "below" is this...
        if graphics_switch_visual_effect_and_powered_position:
            graphics_switch_visual_effect_and_powered = "// no 'visual_effect_and_powered' becuase while it does exist, _visual_effect_and_powered_position is set for livery overrides"
        elif TEMPLATE_ID_FULL in [
            'TPL_03A', 'TPL_03D', 'TPL_03G', 'TPL_16A', 'TPL_17A', 'TPL_17B', 'TPL_17C', 'TPL_17D', 'TPL_17E', 'TPL_32A', 'TPL_32B', 'TPL_32C',
        ] and TEMPLATE_ID_FULL not in ['TPL_02A', 'TPL_02D', 'TPL_02E', 'TPL_02F', 'TPL_42A', 'TPL_42B',
                                       ]:
            graphics_switch_visual_effect_and_powered = f"visual_effect_and_powered: switch_{VEHIDCODE_lcase}_visual_effect;"
        elif TEMPLATE_ID_FULL in ['TPL_02A', 'TPL_02D', 'TPL_02E', 'TPL_02F', 'TPL_42A', 'TPL_42B',
                                  ]:
            graphics_switch_visual_effect_and_powered = f"visual_effect_and_powered: switch_{VEHIDCODE_lcase}_visual_effect_and_powered;"
        else:
            graphics_switch_visual_effect_and_powered = "// no 'visual_effect' or 'visual_effect_and_powered'"

        graphics_switch_articulated_part = f"articulated_part: switch_{VEHIDCODE_lcase}_articulated;" if TEMPLATE_ID_FULL in [
            'TPL_03A', 'TPL_03D', 'TPL_03G', 'TPL_16A', 'TPL_16A', 'TPL_16B', 'TPL_17A', 'TPL_17C', 'TPL_17D', 'TPL_17E', 'TPL_25A', 'TPL_32A', 'TPL_32B', 'TPL_32C',] else "// no 'articulated_part'"
        graphics_switch_length = f"length: switch_{VEHIDCODE_lcase}_length;" if TEMPLATE_ID_FULL in [
            'TPL_03A', 'TPL_03D', 'TPL_03G', 'TPL_16A', 'TPL_16B', 'TPL_17C', 'TPL_17D', 'TPL_17E', 'TPL_32A', 'TPL_32B', 'TPL_32C',] else "// no 'length'"

        # These are for livery overrides
        # Actually I think we only care about middle and cargo -> basically it's to say that if there is something between the two ends
        # ... then it should look like xy so even though 1 and only 1 item has front/back, it's never been called even in legacy code.
        # Also no ending ";" for these on purpose.
        graphics_switch_front_livery = f"switch_{VEHIDCODE_lcase}_front_livery" if TEMPLATE_ID_FULL in [
            'TPL_42A'] else None
        graphics_switch_middle_livery = f"switch_{VEHIDCODE_lcase}_middle_livery" if TEMPLATE_ID_FULL in [
            'TPL_25A', 'TPL_42A',] else None
        graphics_switch_back_livery = f"switch_{VEHIDCODE_lcase}_back_livery" if TEMPLATE_ID_FULL in [
            'TPL_42A'] else None
        graphics_switch_cargo_selection = f"switch_{VEHIDCODE_lcase}_cargo_selection" if TEMPLATE_ID_FULL in ['TPL_02A', 'TPL_02D', 'TPL_02E', 'TPL_02F', 'TPL_02F', 'TPL_04C', 'TPL_04E', 'TPL_04F', 'TPL_04G', 'TPL_04H', 'TPL_04I', 'TPL_04J', 'TPL_04K', 'TPL_04L', 'TPL_04M', 'TPL_04N', 'TPL_04O', 'TPL_04P', 'TPL_04Q',
                                                                                                              ] else None

        # This middle is not the middle above...[we ignore the 3-4 'steam' types that also actually have this because in legacy code i checked and it's not being applied.]
        graphics_spriteset_middle = f"spriteset_{VEHIDCODE_lcase}_middle" if is_true(
            category == 'METRO') and not is_true(row['IS_WAGON_OR_COACH']) else None

        graphics_switch_can_attach = f"can_attach_wagon: switch_can_attach_vehicle;" if not is_true(
            row['IS_WAGON_OR_COACH']) else "// no 'can_attach'"

        graphics_spritegroup_t42b_head_logic = f"spritegroup_{VEHIDCODE_lcase}_engine1_l1" if TEMPLATE_ID_FULL in [
            'TPL_42B'] else None

        graphics_switch_t42b_wagon_logic = f"switch_{VEHIDCODE_lcase}_wagon_logic" if TEMPLATE_ID_FULL in [
            'TPL_42B'] else None

        content = []
        content.append(f"\n{copyright_text}\n\n")
        content.append(
            f"\n// Template: {TEMPLATE_ID_FULL}.\n// Data from: {row['WEB']}\n\n")
        if veh_notes:
            content.append("// Notes:\n")
            for k, v in veh_notes.items():
                content.append(f"//// {k}: {v}\n")
            content.append("\n\n")

        # We need to port some of the random crap from _graphics here else it won't work because we are no longer defining HEAD_CAPACITY as a generic thing.
        if (category in ['DMU', 'EMU', 'WAGON', 'MAGLEVMU'] and row['VEHID_ID_CATEGORY'] not in ['ID_RANGE_CARGOEMU', 'ID_RANGE_CARGODMU']) or category.endswith('RAILBUS'):
            content.append("// Cargo capacity" + "\n")
            content.append(get_expanded_engine_capacity_switch(row))
            content.append(get_expanded_wagon_capacity_switch(row))

        # I've wholly failed to figure out why these two are special in a logical way so i'm just hardcoding them
        if VEHIDCODE_lcase in ['rbd_germany_saxon_det_1_2', 'rbs_south_africa_csar_railmotor']:
            content.append(f"""switch(FEAT_TRAINS, SELF, switch_{VEHIDCODE_lcase}_capacity_position, position_in_vehid_chain % 2) {{
                0: switch_{VEHIDCODE_lcase}_capacity_engine;
                0;
            }}\n\n""")

        content.append(f"item(FEAT_TRAINS, {row['ITEM'].lower()}) {{\n")
        content.append("    property {\n")
        content.append(f"        name: string({row['NAME'].lower()});\n")
        content.append(f"        {get_climates(row)}\n")
        content.append(
            f"        introduction_date: date({int(row['INTRODUCTION_YEAR'])},1,1);\n")
        content.append(
            f"        model_life: {"VEHICLE_NEVER_EXPIRES" if row['MODEL_LIFE'] == "VEHICLE_NEVER_EXPIRES" else int(row['MODEL_LIFE'])};\n")
        content.append(f"        vehicle_life: {int(row['VEHICLE_LIFE'])};\n")
        content.append(f"        retire_early: {RETIRE_EARLY};\n")
        content.append(f"        loading_speed: {ls_logic};\n")
        content.append(f"        cost_factor: {p_cost};\n")
        content.append(f"        running_cost_factor: {r_cost};\n")
        content.append(f"        speed: {int(row['SPEED'])} km/h;\n")
        content.append(f"        power: {int(row['POWER'])} hp;\n")
        content.append(
            f"        cargo_capacity: {int(row['HEAD_CAPACITY'])};\n")
        content.append(f"        weight: {int(row['WEIGHT'])} ton;\n")
        content.append(
            f"        tractive_effort_coefficient: {row['TE_COEFFICIENT']};\n")
        content.append(
            f"        air_drag_coefficient: {AIR_DRAG_COEFFICIENT};\n\n")
        content.append(
            f"        reliability_decay: {RELIABILITY_DECAY};\n")
        content.append(f"        {get_cargo_definitions(row)}\n")
        content.append(
            f"        cargo_age_period: {CARGO_AGE_PERIOD};\n")
        content.append(f"        misc_flags: {misc_logic};\n")
        content.append(f"        refit_cost: {REFIT_COST};\n")
        content.append(
            f"        ai_special_flag: {"AI_FLAG_PASSENGER" if is_true(row['PASSENGER']) else "AI_FLAG_CARGO"};\n")
        content.append(f"        track_type: {track_logic};\n")
        content.append(
            f"        running_cost_base: {RUNNING_COST_BASE};\n")
        content.append(
            f"        engine_class: {'ENGINE_CLASS_' + row['ENGINE_CLASS']};\n")
        content.append(
            f"        visual_effect_and_powered: visual_effect_and_powered({v1}, {row['VISUAL_EFFECT_2']}, {row['VISUAL_EFFECT_3']});\n\n")
        content.append(f"        sprite_id: {SPRITE_ID};\n")
        content.append(f"        dual_headed: {int(row['DUAL_HEADED'])};\n")
        content.append(f"        length: {int(row['LENGTH'])};\n")
        content.append(
            f"        extra_power_per_wagon: {POWER_PER_WAGON};\n")
        content.append(
            f"        bitmask_vehicle_info: {BITMASK_VEHICLE_INFO};\n")

        if is_true(row['IS_WAGON_OR_COACH']):
            pass
        else:
            content.append(f"{get_badges(row)}\n")
        content.append("    }\n")
        # Graphics selection/overrides
        content.append("    graphics {\n")
        cargo_capacity_defined = False
        if TEMPLATE_ID_FULL in ['TPL_32B']:
            content.append(
                f"        cargo_capacity: switch_{VEHIDCODE_lcase}_capacity_position;\n")
            cargo_capacity_defined = True

        content.append(
            f"        purchase: spriteset_{VEHIDCODE_lcase}_purchase;\n")
        if is_true(row['IS_POWERED_UNPOWERED_SUNDRY']):
            # FML.
            purchasetext = "PURCHASETEXT"
            cargodef = "PASSENGER" if is_true(
                row['CARGODEF'].startswith('PASSENGER')) else "CARGO"
            powered_state = "UNPOWERED" if VEHIDCODE_lcase.endswith(
                'unpowered') else "POWERED"
            content.append(
                f"        {purchasetext}MUWAGON{cargodef}{powered_state}\n")
        elif (category in ['DMU', 'EMU', 'WAGON', 'MAGLEVMU'] and row['VEHID_ID_CATEGORY'] not in ['ID_RANGE_CARGODMU', 'ID_RANGE_CARGOEMU']) or category.endswith('RAILBUS'):
            if not cargo_capacity_defined:
                content.append(
                    f"        cargo_capacity: switch_{VEHIDCODE_lcase}_capacity_engine;\n")
        if purchase_cargo_capacity and purchase_cargo_capacity > 0:
            content.append(
                f"        purchase_cargo_capacity: {int(purchase_cargo_capacity)};\n")
        content.append(
            f"        {graphics_switch_visual_effect_and_powered}\n")
        content.append(f"        {graphics_switch_length}\n")
        content.append(f"        {graphics_switch_articulated_part}\n")
        content.append(f"        {graphics_switch_can_attach}\n")

        if not is_true(row['IS_POWERED_UNPOWERED_SUNDRY']):
            content.append(f"        // Add calls to defined switches below\n")
            content.append(
                f"        // RUNNINGCOST_ENGINE_SWITCH_CALL // this is actually blank\n")
            content.append(
                f"        // PURCHASETEXT_SWITCH_CALL // this is actually blank\n")
            pass

        if TEMPLATE_ID_FULL in ['TPL_03A', 'TPL_03D', 'TPL_03F', 'TPL_03G', 'TPL_16A', 'TPL_16B', 'TPL_17A', 'TPL_17B', 'TPL_17C', 'TPL_17D', 'TPL_17E', 'TPL_32A', 'TPL_32B', 'TPL_32C',
                                ]:
            content.append(
                f"        default: switch_{VEHIDCODE_lcase}_position;\n")
        elif TEMPLATE_ID_FULL in [
            'TPL_42B'
        ]:
            content.append(
                f"        default: {graphics_spritegroup_t42b_head_logic};\n")
        elif TEMPLATE_ID_FULL in ['TPL_03A', 'TPL_03B', 'TPL_03C', 'TPL_03E', 'TPL_03G', 'TPL_32B', 'TPL_32C',
                                  ]:
            content.append(
                f"        default: switch_{VEHIDCODE_lcase}_animation;\n")
        elif category in ['DIESELENGINE', 'ELECTRICENGINE', 'MAGLEVSU', 'STEAMENGINE'] \
                or category.endswith('RAILBUS') \
                or is_true(row['IS_POWERED_UNPOWERED_SUNDRY']):
            if TEMPLATE_ID_FULL in ['TPL_02A', 'TPL_02C', 'TPL_02D', 'TPL_02E', 'TPL_02F', 'TPL_42A', 'TPL_42B',
                                    ]:
                content.append(
                    f"        default: switch_{VEHIDCODE_lcase}_reversed;\n")
            else:
                content.append(
                    f"        default: spriteset_{VEHIDCODE_lcase};\n")
        elif category in ['COACH', 'WAGON']:
            if TEMPLATE_ID_FULL in ['TPL_02A', 'TPL_02D', 'TPL_02E', 'TPL_02F', 'TPL_04C', 'TPL_04E', 'TPL_04F', 'TPL_04G', 'TPL_04H', 'TPL_04I', 'TPL_04J', 'TPL_04K', 'TPL_04L', 'TPL_04M', 'TPL_04N', 'TPL_04O', 'TPL_04P', 'TPL_04Q',
                                    ]:
                content.append(
                    f"        default: switch_{VEHIDCODE_lcase}_cargo_selection;\n")
            elif TEMPLATE_ID_FULL in [
                'TPL_04C',
                'TPL_04D',
                'TPL_04J',
                'TPL_04M',
                'TPL_04P',
            ]:
                # Total cluserf.k but box-cars and some tanker-wagons have so-called standard liveries
                # ....with a capital 'S'!
                content.append(
                    f"        default: switch_{VEHIDCODE_lcase}_standard_livery;\n")
            elif TEMPLATE_ID_FULL in [
                'TPL_02F',
                'TPL_04A',
                'TPL_04B',
                'TPL_04R',
            ]:
                content.append(
                    f"        default: switch_{VEHIDCODE_lcase}_livery;\n")
            else:
                # I've lost track of this sh.t by now.
                content.append(f"        default: switch_{VEHIDCODE_lcase};\n")
        elif category in ['DMU', 'EMU', 'METRO', 'MAGLEVMU']:
            content.append(
                f"        default: switch_{VEHIDCODE_lcase}_reversed;\n")
        else:
            pass
        content.append("    }\n")

        # Livery Overrides for MUs - but not actual wagons.
        # TBH I have no idea why powered/unpowered wagons are generally classified as *MUs, rather than COACHes/WAGONs but I'll leave it as-is.
        if (category in ['DMU', 'EMU', 'METRO', 'MAGLEVMU'] and not is_true(row['IS_POWERED_UNPOWERED_SUNDRY'])):

            # At the moment there's just 1 CARGOxMU..
            overrideType = "cargo_" if row['LOADINGSPEED'] == "CARGO" and is_true(
                row['MISC_FLAGS_TRAIN_FLAG_MU']) else ""

            # Unpowered Wagon
            if category in ['METRO']:
                content.append(
                    f"    livery_override (item_mtro_metro_{overrideType}wagon_unpowered) {{\n")
            else:
                content.append(
                    f"    livery_override (item_mu_mu_{overrideType}wagon_unpowered) {{\n")
            content.append(f"        loading_speed: {ls_logic};\n")
            content.append(
                f"        running_cost_factor: int({int(row['SPEED'])}/10);\n")
            content.append(f"        weight: int({int(row['WEIGHT'])}*1/2);\n")
            if graphics_switch_visual_effect_and_powered_position:
                content.append(
                    f"        {graphics_switch_visual_effect_and_powered_position}\n")
            content.append(
                f"        cargo_capacity: {int(row['WAGON_CAPACITY'])};\n")
            content.append(f"        length: {int(row['WAGON_LENGTH'])};\n")
            content.append(f"        default: {graphics_switch_middle_livery if graphics_switch_middle_livery
                                               else graphics_switch_cargo_selection if graphics_switch_cargo_selection
                                               else graphics_spriteset_middle if graphics_spriteset_middle
                                               else graphics_switch_t42b_wagon_logic if graphics_switch_t42b_wagon_logic else None};\n")
            content.append(f"    }}\n")

            # Powered Wagon - Replicating complex RC eval
            p_sqrt = nml_sqrt(row['POWER'])
            rc_powered = (int(row['SPEED'])/10) + (p_sqrt/10) + \
                (row['TE_COEFFICIENT'] * row['WEIGHT'])

            if category in ['METRO']:
                content.append(
                    f"    livery_override (item_mtro_metro_{overrideType}wagon_powered) {{\n")
            else:
                content.append(
                    f"    livery_override (item_mu_mu_{overrideType}wagon_powered) {{\n")
            content.append(f"        loading_speed: {ls_logic};\n")
            content.append(
                f"        running_cost_factor: (round({rc_powered}));\n")
            content.append(f"        power: int({int(row['POWER'])}*1/2);\n")
            content.append(f"        weight: int({int(row['WEIGHT'])}*3/4);\n")
            if graphics_switch_visual_effect_and_powered_position:
                content.append(
                    f"        {graphics_switch_visual_effect_and_powered_position}\n")
            content.append(
                f"        cargo_capacity: {int(row['WAGON_CAPACITY'])};\n")
            content.append(
                f"        tractive_effort_coefficient: int({row['TE_COEFFICIENT']}*10*{int(row['WEIGHT'])});\n")
            content.append(f"        length: {int(row['WAGON_LENGTH'])};\n")
            content.append(f"        default: {graphics_switch_middle_livery if graphics_switch_middle_livery
                                               else graphics_switch_cargo_selection if graphics_switch_cargo_selection
                                               else graphics_spriteset_middle if graphics_spriteset_middle
                                               else graphics_switch_t42b_wagon_logic if graphics_switch_t42b_wagon_logic else None};\n")
            content.append(f"    }}\n")

        content.append("}\n")

        # Save Logic
        rel_folder = str(row['SAVE_TO']).replace('\\', '/')
        abs_folder = os.path.normpath(os.path.join(project_root, rel_folder))
        os.makedirs(abs_folder, exist_ok=True)
        with open(os.path.join(abs_folder, f"{row['FILENAMES_EXPECTED']}_item.pnml"), 'w', encoding='utf-8') as f:
            f.writelines(content)

    # Save as CSV - this takes quite a few seconds.
    print("---- Saving CSV File ----")
    df_master.to_csv(os.path.join(
        script_dir, 'vehicle_report.csv'), index=False)
    print("---- Saving CSV File Complete ----")


if __name__ == "__main__":
    generate_unified_items()
    print("--- Unified Item File Generation Complete ---")
